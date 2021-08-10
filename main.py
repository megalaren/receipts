import os
import sys

from fpdf import FPDF
from openpyxl import load_workbook

from models import *


BASE_DIRECTORY = os.getcwd()
DATE_FORMAT = '%d.%m.%Y'
NOW = dt.datetime.today()
FILE_NAME_PEOPLE = 'people.xlsx'
SHEET_NAME_PEOPLE = 'Люди'
SHEET_NAME_CUSTOMIZATION = 'Настройки'
PATH_TO_PEOPLE = os.path.join(BASE_DIRECTORY, FILE_NAME_PEOPLE)

HEIGHT = 4
PAGE_HEIGHT = 275
PAGE_WIDTH = 180
TOP_MARGIN = 10
LEFT_MARGIN = 15
BOTTOM_MARGIN = 10
FONT_SIZE = 8


def resource_path(relative):
    """Для получения доступа к ресурсам.

    Нужно для создания исполняемого файла.
    Описание здесь - https://irwinkwan.com/tag/pyinstaller/"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(relative)


def get_height_multi_cell(font_file, font_size, text: str) -> int:
    """Возвращает высоту, которую займет текст в multi_cell."""
    pdf = FPDF()
    pdf.add_font('Current_font', '', font_file, uni=True)
    pdf.set_font('Current_font', '', font_size)
    pdf.add_page()
    pdf.set_y(0)
    start_y = pdf.get_y()
    pdf.multi_cell(PAGE_WIDTH, HEIGHT, text, align='L')
    end_y = pdf.get_y()
    return int(end_y - start_y)


def main():
    try:
        wb = load_workbook(PATH_TO_PEOPLE, read_only=True, data_only=True)
    except FileNotFoundError:
        file_people = FILE_NAME_PEOPLE.split('.')
        raise FileNotFoundError(f'Проверьте что excel файл называется "{file_people[0]}" '
                                f'и имеет расширение "{file_people[1]}"')
    font_filename = 'DejaVuSansCondensed.ttf'
    font_file = resource_path(font_filename)
    pdf = FPDF()
    pdf.add_font('DejaVu', '', font_file, uni=True)
    pdf.set_top_margin(TOP_MARGIN)
    pdf.set_left_margin(LEFT_MARGIN)
    pdf.set_auto_page_break(True, BOTTOM_MARGIN)

    customization = get_customization(wb)
    people = get_people(wb).values()
    pdf.add_page()
    remaining_height = PAGE_HEIGHT
    height_bottom_line = get_height_multi_cell(font_file, FONT_SIZE, customization['Последняя строка'])

    for i, person in enumerate(people):
        # если нет задолженности - пропускаем
        if person.get_sum_debt() == 0:
            continue
        # определим, надо ли добавить новую страницу
        person_height = person.get_height(height_bottom_line)
        if remaining_height == PAGE_HEIGHT:
            remaining_height -= person_height
        elif remaining_height >= person_height + HEIGHT * 2:
            pdf.cell(PAGE_WIDTH, HEIGHT, '_ ' * 86, ln=1, align='C')
            pdf.cell(PAGE_WIDTH, HEIGHT, ln=1)
            remaining_height -= person_height + HEIGHT * 2
        else:
            pdf.add_page()
            remaining_height = PAGE_HEIGHT - person_height

        pdf.set_font('DejaVu', '', FONT_SIZE)
        pdf.cell(PAGE_WIDTH, HEIGHT, str(customization['Название']), 0, 1, 'C', )
        pdf.cell(PAGE_WIDTH, HEIGHT, str(customization['Заголовок']), 0, 1, 'C', )
        date_of_receipt_formation = customization['Дата формирования'].strftime(DATE_FORMAT)
        pdf.cell(PAGE_WIDTH, HEIGHT, f'Сформировано на {date_of_receipt_formation}г.', 0, 1, 'L', )
        pdf.cell(PAGE_WIDTH, HEIGHT * 4, '', 1, 0, 'C', )
        pdf.cell(-PAGE_WIDTH)
        pdf.cell(PAGE_WIDTH, HEIGHT, str(customization['Получатель платежа 1 строка']), 0, 1, 'L', )
        pdf.cell(PAGE_WIDTH, HEIGHT, str(customization['Получатель платежа 2 строка']), 0, 1, 'L', )
        pdf.cell(PAGE_WIDTH, HEIGHT, str(customization['Получатель платежа 3 строка']), 0, 1, 'L', )
        pdf.cell(PAGE_WIDTH, HEIGHT, str(customization['Получатель платежа 4 строка']), 0, 1, 'L', )
        pdf.cell(PAGE_WIDTH, HEIGHT, f'Плательщик: {person.name} ({person.phone})', 0, 1, 'L', )
        pdf.cell(PAGE_WIDTH, HEIGHT, f'Адрес: {person.address}', 0, 1, 'L', )
        pdf.cell(PAGE_WIDTH, HEIGHT, '', 0, 1, 'L', )

        for box in person.boxes:
            pdf.cell(PAGE_WIDTH, HEIGHT, f'Гараж № {box.number}', 0, 1, 'L', )
            pdf.cell(int(PAGE_WIDTH / 3), HEIGHT, 'Начислено членских взносов', 1, 0, 'C', )
            pdf.cell(int(PAGE_WIDTH / 3), HEIGHT, 'Оплачено', 1, 0, 'C', )
            pdf.cell(int(PAGE_WIDTH / 3), HEIGHT, 'Задолженность', 1, 1, 'C', )
            pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, 'Год', 1, 0, 'C', )
            pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, 'Сумма', 1, 0, 'C', )
            pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, 'Дата', 1, 0, 'C', )
            pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, 'Сумма', 1, 0, 'C', )
            pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, 'Срок оплаты', 1, 0, 'C', )
            pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, 'Сумма', 1, 1, 'C', )

            for accrual in box.accruals:
                pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, str(accrual.year), 1, 0, 'C', )
                pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, str(accrual.accrued), 1, 0, 'C', )
                pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, accrual.payment_date, 1, 0, 'C', )

                if accrual.payment_amount == 0:
                    payment_amount = ''
                else:
                    payment_amount = str(accrual.payment_amount)
                pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, payment_amount, 1, 0, 'C', )
                due_date = accrual.due_date.strftime(DATE_FORMAT)
                pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, due_date, 1, 0, 'C', )
                pdf.cell(int(PAGE_WIDTH / 6), HEIGHT, str(accrual.debt), 1, 1, 'C', )
            pdf.cell(20, h=HEIGHT, ln=1)

        pdf.cell(PAGE_WIDTH, HEIGHT, f'Всего начислено: {person.get_sum_accrued()} руб.', 0, 1, 'R', )
        sum_overdue_debt = person.get_sum_overdue_debt(customization['Дата формирования'])
        pdf.cell(PAGE_WIDTH, HEIGHT, f'Всего просроченная задолженность: {sum_overdue_debt} руб.', 0, 1, 'R', )
        sum_debt = str(person.get_sum_debt())
        pdf.cell(PAGE_WIDTH, HEIGHT, f'К уплате по сроку {person.get_last_due_date()}г.: {sum_debt} руб', 0, 1, 'R', )
        pdf.cell(20, h=HEIGHT, ln=1)
        pdf.multi_cell(PAGE_WIDTH, HEIGHT, customization['Последняя строка'], border=0, align='L')

    try:
        pdf.output(os.path.join(BASE_DIRECTORY, 'Квитанции.pdf'), 'F')
    except PermissionError:
        raise PermissionError('Проверьте, что закрыли pdf-файл "Квитанции" перед созданием нового')


def get_people(wb):
    first_row = 3  # первая строка с людьми
    try:
        ws = wb[SHEET_NAME_PEOPLE]
    except KeyError:
        raise KeyError(f'Проверьте, что в excel файле лист со списком людей называется "{SHEET_NAME_PEOPLE}"')
    rows = ws.values
    try:
        ind_phone = next(rows).index('Телефон')
        headlines = next(rows)  # заголовки столбцов
        ind_name = headlines.index('собственника')
        ind_address = headlines.index('Жительства')
        ind_year = headlines.index('Год')
        ind_accrued = ind_year + 1
        ind_payment_date = headlines.index('Дата')
        ind_payment_amount = ind_payment_date + 1
        ind_due_date = headlines.index('Срок оплаты')
        ind_debt = ind_due_date + 1
    except ValueError:
        raise ValueError(f'Проверьте названия столбцов на листе "{SHEET_NAME_PEOPLE}"')

    people = {}
    name = ''

    for i, row in enumerate(rows):
        ind_row = i + 3
        if ind_row < first_row:
            continue
        if row[0] is not None:
            try:
                number = int(row[0])
            except (ValueError, TypeError):
                raise ValueError(f'В строке {ind_row} неправильный формат номера бокса')
            box = Box(number=number)
            try:
                name = row[ind_name].title()
            except (ValueError, TypeError, AttributeError):
                raise ValueError(f'В строке {ind_row} неправильный формат ФИО')
            if name not in people:
                try:
                    phone = str(row[ind_phone]).replace('\n', ' ')
                except (ValueError, TypeError, AttributeError):
                    if row[ind_phone] is None:
                        phone = ''
                    else:
                        raise ValueError(f'В строке {ind_row} неправильный формат номера телефона')
                try:
                    address = str(row[ind_address])
                except (ValueError, TypeError):
                    if row[ind_address] is None:
                        address = ''
                    else:
                        raise ValueError(f'В строке {ind_row} неправильный формат адреса')
                person = People(name=name, address=address, phone=phone)
                people[name] = person
            people[name].boxes.append(box)

        year = row[ind_year]
        try:
            year = int(year)
        except (ValueError, TypeError):
            continue

        accrued = row[ind_accrued]
        try:
            accrued = int(float(accrued))
        except (ValueError, TypeError):
            continue

        payment_date = row[ind_payment_date]
        if payment_date is None:
            payment_date = ''
        elif type(payment_date) == dt.datetime:
            payment_date = payment_date.strftime(DATE_FORMAT)
        else:
            try:
                payment_date = str(payment_date)
            except (ValueError, TypeError):
                raise ValueError(f'В строке {ind_row} неправильный формат даты оплаты')

        payment_amount = row[ind_payment_amount]
        try:
            payment_amount = int(float(payment_amount))
        except (ValueError, TypeError):
            if payment_amount is None or (type(payment_amount) == str and payment_amount.isspace()):  # noqa
                payment_amount = 0
            else:
                raise ValueError(f'В строке {ind_row} неправильный формат суммы оплаты')

        due_date = row[ind_due_date]
        if type(due_date) != dt.datetime:
            try:
                due_date = dt.datetime.strptime(due_date, DATE_FORMAT).date()
            except (ValueError, TypeError, AttributeError):
                raise ValueError(f'В строке {ind_row} неправильный формат даты срока оплаты')

        debt = row[ind_debt]
        try:
            debt = int(float(debt))
        except (ValueError, TypeError):
            if debt is None or (type(debt) == str and debt.isspace()):  # noqa
                debt = 0
            else:
                raise ValueError(f'В строке {ind_row} неправильный формат суммы задолженности')

        accrual = Accrual(year=year, accrued=accrued, payment_date=payment_date,
                          payment_amount=payment_amount, due_date=due_date, debt=debt)
        people[name].boxes[-1].accruals.append(accrual)

    return people


def get_customization(wb):
    keys = ['Название', 'Заголовок', 'Дата формирования', 'Получатель платежа 1 строка',
            'Получатель платежа 2 строка', 'Получатель платежа 3 строка', 'Получатель платежа 4 строка',
            'Последняя строка', ]
    try:
        ws = wb[SHEET_NAME_CUSTOMIZATION]
    except KeyError:
        raise KeyError(f'Проверьте, что в excel файле лист с настройками называется "{SHEET_NAME_CUSTOMIZATION}"')
    rows = ws.values
    customization = {}
    for row in rows:
        key, value = row[0], row[1]
        if key not in keys:
            continue
        if key == 'Дата формирования':
            if type(value) != dt.datetime:
                try:
                    value = dt.datetime.strptime(value, DATE_FORMAT).date()
                except Exception:
                    raise ValueError(f'Неправильный формат поля "{key}" на листе настроек')
        elif key == 'Последняя строка':
            pass
        customization[key] = value

    for key in keys:
        if key not in customization:
            raise KeyError(f'Проверьте, что на листе настроек есть поле "{key}"')
    return customization


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print('В ходе работы программы возникла ошибка:')
        print(e)
        input('Нажмите Enter, чтобы выйти')
    sys.exit()
